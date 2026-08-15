from django.shortcuts import render, redirect, get_object_or_404
from django.db.models import Q
import json
import logging
from django.conf import settings
from django.http import JsonResponse
from django.utils import timezone
from django.views.decorators.http import require_POST, require_GET
from .models import (
    Product, Allergen, MealCategory, MealTime,
    CalorieCategory, CalorieCategoryMeal, Material,
    SwapGroup, SwapItem,
    ClaudeRationGroup, ClaudeRation, ClaudeRationSlot,
    SLOT_TYPES, SLOT_ORDER, SLOT_LABELS, RATION_SLOT_TYPES,
)
from .roles import editor_required

logger = logging.getLogger(__name__)

# ── Каталог продуктов ─────────────────────────────────────────────────────────

def get_filtered_products(request):
    """Применяет все GET-фильтры каталога.
    Возвращает (queryset, filters_dict, selected_meal_categories, selected_allergens)."""
    products = Product.objects.prefetch_related("meal_categories", "allergens").all()

    search                   = request.GET.get("search", "").strip()
    selected_meal_categories = request.GET.getlist("meal_category")
    selected_allergens       = request.GET.getlist("allergen")
    cost_min      = request.GET.get("cost_min", "")
    cost_max      = request.GET.get("cost_max", "")
    kcal_min      = request.GET.get("kcal_min", "")
    kcal_max      = request.GET.get("kcal_max", "")
    protein_min   = request.GET.get("protein_min", "")
    protein_max   = request.GET.get("protein_max", "")
    fat_min       = request.GET.get("fat_min", "")
    fat_max       = request.GET.get("fat_max", "")
    carbs_min     = request.GET.get("carbs_min", "")
    carbs_max     = request.GET.get("carbs_max", "")
    kcal_s_min    = request.GET.get("kcal_s_min", "")
    kcal_s_max    = request.GET.get("kcal_s_max", "")
    protein_s_min = request.GET.get("protein_s_min", "")
    protein_s_max = request.GET.get("protein_s_max", "")
    fat_s_min     = request.GET.get("fat_s_min", "")
    fat_s_max     = request.GET.get("fat_s_max", "")
    carbs_s_min   = request.GET.get("carbs_s_min", "")
    carbs_s_max   = request.GET.get("carbs_s_max", "")
    packing       = request.GET.get("packing", "").strip()
    unused        = request.GET.get("unused", "")
    used          = request.GET.get("used", "")
    sort_by       = request.GET.get("sort", "name")
    sort_dir      = request.GET.get("dir", "asc")

    if search:
        products = products.filter(
            Q(name__icontains=search)
            | Q(composition__icontains=search)
            | Q(composition_clean__icontains=search)
        )
    if selected_meal_categories:
        products = products.filter(meal_categories__key__in=selected_meal_categories).distinct()
    if selected_allergens:
        products = products.filter(allergens__pk__in=selected_allergens).distinct()
    if cost_min:       products = products.filter(cost__gte=cost_min)
    if cost_max:       products = products.filter(cost__lte=cost_max)
    if kcal_min:       products = products.filter(kcal_per_100__gte=kcal_min)
    if kcal_max:       products = products.filter(kcal_per_100__lte=kcal_max)
    if protein_min:    products = products.filter(protein__gte=protein_min)
    if protein_max:    products = products.filter(protein__lte=protein_max)
    if fat_min:        products = products.filter(fat__gte=fat_min)
    if fat_max:        products = products.filter(fat__lte=fat_max)
    if carbs_min:      products = products.filter(carbs__gte=carbs_min)
    if carbs_max:      products = products.filter(carbs__lte=carbs_max)
    if kcal_s_min:     products = products.filter(kcal_per_serving__gte=kcal_s_min)
    if kcal_s_max:     products = products.filter(kcal_per_serving__lte=kcal_s_max)
    if protein_s_min:  products = products.filter(protein_per_serving__gte=protein_s_min)
    if protein_s_max:  products = products.filter(protein_per_serving__lte=protein_s_max)
    if fat_s_min:      products = products.filter(fat_per_serving__gte=fat_s_min)
    if fat_s_max:      products = products.filter(fat_per_serving__lte=fat_s_max)
    if carbs_s_min:    products = products.filter(carbs_per_serving__gte=carbs_s_min)
    if carbs_s_max:    products = products.filter(carbs_per_serving__lte=carbs_s_max)
    if packing:        products = products.filter(packing__icontains=packing)
    # Использование блюда = стоит в рационе Claude или в подгруппе на замену.
    # Старые рационы (Ration) не считаем — вкладку убрали, данные заморожены.
    # Фильтры взаимоисключающие: если каким-то образом пришли оба — не фильтруем.
    if unused and not used:
        products = products.filter(claude_ration_slots__isnull=True, swap_items__isnull=True)
    elif used and not unused:
        products = products.filter(
            Q(claude_ration_slots__isnull=False) | Q(swap_items__isnull=False)
        ).distinct()

    sort_map = {
        "name": "name", "article": "article", "net_weight": "net_weight", "cost": "cost",
        "kcal": "kcal_per_100", "protein": "protein", "fat": "fat", "carbs": "carbs",
        "kcal_s": "kcal_per_serving", "protein_s": "protein_per_serving",
        "fat_s": "fat_per_serving", "carbs_s": "carbs_per_serving",
    }
    sort_field = sort_map.get(sort_by, "name")
    if sort_dir == "desc":
        sort_field = f"-{sort_field}"
    products = products.order_by(sort_field)

    filters = {
        "search": search, "cost_min": cost_min, "cost_max": cost_max,
        "kcal_min": kcal_min, "kcal_max": kcal_max,
        "protein_min": protein_min, "protein_max": protein_max,
        "fat_min": fat_min, "fat_max": fat_max,
        "carbs_min": carbs_min, "carbs_max": carbs_max,
        "kcal_s_min": kcal_s_min, "kcal_s_max": kcal_s_max,
        "protein_s_min": protein_s_min, "protein_s_max": protein_s_max,
        "fat_s_min": fat_s_min, "fat_s_max": fat_s_max,
        "carbs_s_min": carbs_s_min, "carbs_s_max": carbs_s_max,
        "packing": packing, "unused": unused, "used": used,
        "sort": sort_by, "dir": sort_dir,
    }
    return products, filters, selected_meal_categories, selected_allergens


def product_list(request):
    products, filters, selected_meal_categories, selected_allergens = get_filtered_products(request)
    all_meal_categories = MealCategory.objects.all()
    all_allergens = Allergen.objects.all()

    total = products.count()
    packings = (
        Product.objects.values_list("packing", flat=True)
        .distinct().exclude(packing__isnull=True).exclude(packing="").order_by("packing")
    )
    selected_allergen_names = list(
        Allergen.objects.filter(pk__in=selected_allergens).values_list("name", flat=True)
    ) if selected_allergens else []

    return render(request, "myapp/product_list.html", {
        "products": products, "total": total,
        "all_meal_categories": all_meal_categories,
        "all_allergens": all_allergens, "packings": packings,
        "selected_meal_categories": selected_meal_categories,
        "selected_allergens": selected_allergens,
        "selected_allergen_names": selected_allergen_names,
        "slot_labels": SLOT_LABELS,
        "filters": filters,
    })


def product_export(request):
    """Экспорт каталога в Excel с учётом применённых фильтров."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from django.http import HttpResponse
    from urllib.parse import quote

    products, _, _, _ = get_filtered_products(request)
    products = products.prefetch_related("meal_categories")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Каталог"

    green_dark  = PatternFill("solid", fgColor="2E7D32")
    green_light = PatternFill("solid", fgColor="E8F5E0")
    title_font  = Font(bold=True, color="FFFFFF", size=14)
    bold        = Font(bold=True, size=11)
    thin        = Side(style="thin", color="DDDDDD")
    border      = Border(left=thin, right=thin, top=thin, bottom=thin)
    center      = Alignment(horizontal="center", vertical="center")
    left_wrap   = Alignment(horizontal="left", vertical="center", wrap_text=True)

    headers = [
        "Категория", "Наименование", "Артикул", "Кратность", "Масса, г", "Себест., ₸",
        "Белки/100г", "Жиры/100г", "Углев./100г", "Ккал/100г", "КДж/100г",
        "Белки (порц)", "Жиры (порц)", "Углев. (порц)", "Ккал (порц)", "КДж (порц)",
        "Состав",
    ]
    ncols = len(headers)

    def num(val):
        if val is None:
            return None
        try:
            return round(float(val), 2)
        except (TypeError, ValueError):
            return None

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    c = ws.cell(row=1, column=1, value="Каталог блюд")
    c.font = title_font
    c.fill = green_dark
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 26

    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=2, column=col, value=h)
        c.font = bold
        c.fill = green_light
        c.alignment = center
        c.border = border

    row = 3
    for p in products:
        cats = ", ".join(str(mc) for mc in p.meal_categories.all())
        weight_g = num(p.net_weight * 1000) if p.net_weight is not None else None
        values = [
            cats, p.name, p.article or "", p.packing or "", weight_g, num(p.cost),
            num(p.protein), num(p.fat), num(p.carbs), num(p.kcal_per_100), num(p.kj_per_100),
            num(p.protein_per_serving), num(p.fat_per_serving), num(p.carbs_per_serving),
            num(p.kcal_per_serving), num(p.kj_per_serving),
            p.composition_clean or "",
        ]
        for col, val in enumerate(values, start=1):
            c = ws.cell(row=row, column=col, value=val)
            c.border = border
            if col in (1, 2, 17):
                c.alignment = left_wrap
            else:
                c.alignment = center
        row += 1

    widths = [18, 32, 11, 12, 9, 10, 10, 9, 11, 10, 10, 11, 10, 12, 11, 11, 40]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A3"

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = "attachment; filename*=UTF-8''" + quote("Каталог.xlsx")
    wb.save(response)
    return response


def product_detail(request, pk):
    product = get_object_or_404(Product, pk=pk)
    return render(request, "myapp/product_detail.html", {"product": product})


@require_POST
def product_set_sale_price(request, pk):
    """Сохраняет «Цену продажи с ФЗ» блюда из каталога (колонка в таблице
    или модалка). Введённая цена фиксируется как ручная и больше не
    пересчитывается автоматически; пустое поле снимает ручной режим и
    возвращает цену с целевой наценкой. Наценка в БД не хранится — считается
    и возвращается в ответе, чтобы обновить строку без перезагрузки."""
    from decimal import Decimal, InvalidOperation
    product = get_object_or_404(Product, pk=pk)
    raw = (request.POST.get("sale_price") or "").strip().replace(",", ".").replace(" ", "")
    if raw == "":
        # Возврат к автоцене: save() подставит её сам, если есть себестоимость
        product.sale_price = None
        product.sale_price_manual = False
    else:
        try:
            val = Decimal(raw)
        except (InvalidOperation, ValueError):
            return JsonResponse({"ok": False, "message": "Некорректное число"}, status=400)
        if val < 0:
            return JsonResponse({"ok": False, "message": "Цена не может быть отрицательной"}, status=400)
        product.sale_price = val
        product.sale_price_manual = True
    product.save(update_fields=["sale_price", "sale_price_manual"])
    return JsonResponse({
        "ok": True,
        "sale_price": float(product.sale_price) if product.sale_price is not None else None,
        "markup":      float(product.markup) if product.markup is not None else None,
        "markup_pct":  round(product.markup_pct, 1) if product.markup_pct is not None else None,
        "markup_low":  product.markup_is_low,
        "manual":      product.sale_price_manual,
    })


# ── Вспомогательные константы ─────────────────────────────────────────────────

SLOT_ICONS = {
    'breakfast': '🌅',
    'hot_400': '🍽️',  'hot_500': '🍽️',
    'soup': '🥣',
    'salad': '🥗',
    'dessert': '🍰',
    'smoothie': '🥤',  'sandwich': '🥪',
    'extra': '🛒',
}

SLOT_COLORS = {
    'breakfast': 'orange',
    'hot_400': 'green',  'hot_500': 'green',
    'soup': 'blue',
    'salad': 'teal',
    'dessert': 'pink',
    'smoothie': 'purple',  'sandwich': 'amber',
    'extra': 'green',
}


def _product_picker_dict(p, with_category=False):
    """Данные блюда для пикера в редакторе рациона (КБЖУ на порцию и на 100 г,
    себестоимость, аллергены) — для фильтров и отображения."""
    d = {
        "id": p.pk, "name": p.name, "article": p.article or "",
        # на порцию
        "kcal":    float(p.kcal_per_serving or 0),
        "protein": float(p.protein_per_serving or 0),
        "fat":     float(p.fat_per_serving or 0),
        "carbs":   float(p.carbs_per_serving or 0),
        # на 100 г
        "kcal100":    float(p.kcal_per_100 or 0),
        "protein100": float(p.protein or 0),
        "fat100":     float(p.fat or 0),
        "carbs100":   float(p.carbs or 0),
        "cost":      float(p.cost or 0),
        "allergens": [a.name for a in p.allergens.all()],
        "photo":     p.photo.url if p.photo else "",
    }
    if with_category:
        d["category"] = " / ".join(str(c) for c in p.meal_categories.all())
    return d


# ── Калоражи ──────────────────────────────────────────────────────────────────

def _calorie_meals_preview():
    """{ккал: [{name, icon}, …]} — превью приёмов пищи калоража для модала
    создания рациона (обычного и Claude)."""
    preview = {}
    for category in CalorieCategory.objects.prefetch_related("meals__meal_time"):
        preview[category.kcal] = [
            {"name": mt.name, "icon": mt.icon}
            for mt in category.meal_times()
        ]
    return preview


def _seed_slots_from_calorie_category(ration, kcal_category, slot_model):
    """Создаёт слоты нового рациона по приёмам пищи его калоража."""
    category = CalorieCategory.norm_for(kcal_category)
    if not category:
        return
    for i, meal_time in enumerate(category.meal_times()):
        slot_model.objects.create(ration=ration, meal_time=meal_time, order=i)


def _parse_calorie_form(request):
    """Разбирает форму калоража. Возвращает (поля, id приёмов пищи) или (None, None)
    если калорийность не задана."""
    def _int(name, default=0):
        try:
            return max(0, int(request.POST.get(name, default) or default))
        except (TypeError, ValueError):
            return default

    def _range(prefix, fallback_lo=0, fallback_hi=0):
        """Пара «от, до». Если пользователь перепутал поля местами — меняем их
        обратно, чтобы диапазон всегда был корректным."""
        lo = _int(f"{prefix}_min", fallback_lo)
        hi = _int(f"{prefix}_max", fallback_hi)
        return (hi, lo) if hi < lo else (lo, hi)

    kcal = _int("kcal")
    if not kcal:
        return None, None

    kcal_min, kcal_max = _range("kcal", max(0, kcal - 50), kcal + 50)
    protein_min, protein_max = _range("protein")
    fat_min, fat_max = _range("fat")
    carbs_min, carbs_max = _range("carbs")

    fields = {
        "name": request.POST.get("name", "").strip() or f"{kcal} ккал",
        "kcal": kcal,
        "kcal_min": kcal_min,
        "kcal_max": kcal_max,
        "protein_min": protein_min,
        "protein_max": protein_max,
        "fat_min": fat_min,
        "fat_max": fat_max,
        "carbs_min": carbs_min,
        "carbs_max": carbs_max,
    }
    meal_time_ids = [int(i) for i in request.POST.getlist("meal_times") if str(i).isdigit()]
    return fields, meal_time_ids


def _save_calorie_meals(category, meal_time_ids):
    """Переписывает набор приёмов пищи калоража, сохраняя порядок MealTime."""
    category.meals.all().delete()
    ordered = MealTime.objects.filter(pk__in=meal_time_ids).order_by("order", "name")
    for i, meal_time in enumerate(ordered):
        CalorieCategoryMeal.objects.create(category=category, meal_time=meal_time, order=i)


@editor_required
def calorie_list(request):
    """Вкладка «Калоражи» — только для редакторов."""
    categories = list(
        CalorieCategory.objects.prefetch_related("meals__meal_time")
    )

    # Сколько рационов уже используют калораж — показываем при удалении
    usage = {}
    for c in categories:
        usage[c.pk] = ClaudeRation.objects.filter(kcal_category=c.kcal).count()

    categories_data = [
        {"category": c, "meal_times": c.meal_times(), "usage": usage[c.pk]}
        for c in categories
    ]

    # Данные для модала редактирования — заполняются из JS
    categories_json = [
        {
            "id": c.pk,
            "name": c.name,
            "kcal": c.kcal,
            "kcal_min": c.kcal_min,
            "kcal_max": c.kcal_max,
            "protein_min": c.protein_min,
            "protein_max": c.protein_max,
            "fat_min": c.fat_min,
            "fat_max": c.fat_max,
            "carbs_min": c.carbs_min,
            "carbs_max": c.carbs_max,
            "meal_time_ids": [m.meal_time_id for m in c.meals.all()],
        }
        for c in categories
    ]

    return render(request, "myapp/calorie_list.html", {
        "categories_data": categories_data,
        "all_meal_times": list(MealTime.objects.order_by("order", "name")),
        "categories_json": json.dumps(categories_json, ensure_ascii=False),
    })


@editor_required
def calorie_create(request):
    if request.method == "POST":
        fields, meal_time_ids = _parse_calorie_form(request)
        if fields and not CalorieCategory.objects.filter(kcal=fields["kcal"]).exists():
            fields["order"] = CalorieCategory.objects.count()
            category = CalorieCategory.objects.create(**fields)
            _save_calorie_meals(category, meal_time_ids)
    return redirect("calorie_list")


@editor_required
def calorie_edit(request, pk):
    category = get_object_or_404(CalorieCategory, pk=pk)
    if request.method == "POST":
        fields, meal_time_ids = _parse_calorie_form(request)
        if fields:
            taken = CalorieCategory.objects.filter(kcal=fields["kcal"]).exclude(pk=category.pk)
            if not taken.exists():
                old_kcal = category.kcal
                for name, value in fields.items():
                    setattr(category, name, value)
                category.save()
                _save_calorie_meals(category, meal_time_ids)
                # Рационы ссылаются на калораж числом ккал — переносим их следом,
                # иначе они остались бы без нормы КБЖУ.
                if old_kcal != category.kcal:
                    ClaudeRation.objects.filter(kcal_category=old_kcal).update(kcal_category=category.kcal)
    return redirect("calorie_list")


@editor_required
def calorie_delete(request, pk):
    category = get_object_or_404(CalorieCategory, pk=pk)
    if request.method == "POST":
        category.delete()
    return redirect("calorie_list")


# ── Блюда на замену ───────────────────────────────────────────────────────────

def swap_list(request):
    groups = list(
        SwapGroup.objects.prefetch_related("items__product__allergens").all()
    )
    groups_data = []
    for g in groups:
        items = []
        for it in g.items.all():
            p = it.product
            items.append({
                "item": it,
                "product": p,
            })
        groups_data.append({"group": g, "items": items, "count": len(items)})

    # Все блюда для модала выбора
    all_products = (
        Product.objects.prefetch_related("meal_categories").order_by("name")
    )
    all_products_json = [
        {
            "id": p.pk, "name": p.name, "article": p.article or "",
            "category": " / ".join(str(c) for c in p.meal_categories.all()),
            "kcal":    float(p.kcal_per_serving or 0),
            "protein": float(p.protein_per_serving or 0),
            "fat":     float(p.fat_per_serving or 0),
            "carbs":   float(p.carbs_per_serving or 0),
            "cost":    float(p.cost or 0),
            "photo":   p.photo.url if p.photo else "",
        }
        for p in all_products
    ]

    return render(request, "myapp/swap_list.html", {
        "groups_data": groups_data,
        "all_products_json": json.dumps(all_products_json, ensure_ascii=False),
    })


def swap_group_create(request):
    if request.method == "POST":
        name = request.POST.get("name", "").strip()
        if name:
            last = SwapGroup.objects.count()
            SwapGroup.objects.create(name=name, order=last)
    return redirect("swap_list")


def swap_group_edit(request, group_pk):
    group = get_object_or_404(SwapGroup, pk=group_pk)
    if request.method == "POST":
        name = request.POST.get("name", "").strip()
        if name:
            group.name = name
            group.save(update_fields=["name"])
    return redirect("swap_list")


def swap_group_delete(request, group_pk):
    group = get_object_or_404(SwapGroup, pk=group_pk)
    if request.method == "POST":
        group.delete()
    return redirect("swap_list")


@require_POST
def swap_item_add(request, group_pk):
    group = get_object_or_404(SwapGroup, pk=group_pk)
    product_id = request.POST.get("product_id")
    if product_id:
        product = get_object_or_404(Product, pk=product_id)
        # Не добавляем дубль одного и того же блюда в подгруппу
        if not SwapItem.objects.filter(swap_group=group, product=product).exists():
            last = group.items.count()
            SwapItem.objects.create(swap_group=group, product=product, order=last)
    return redirect("swap_list")


@require_POST
def swap_item_remove(request, item_pk):
    item = get_object_or_404(SwapItem, pk=item_pk)
    item.delete()
    return redirect("swap_list")


def swap_group_export(request, group_pk):
    """Экспорт одной подгруппы блюд на замену в Excel."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from django.http import HttpResponse
    from urllib.parse import quote

    group = get_object_or_404(SwapGroup, pk=group_pk)
    items = list(group.items.select_related("product").prefetch_related("product__allergens"))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Блюда на замену"

    green_dark  = PatternFill("solid", fgColor="2E7D32")
    green_light = PatternFill("solid", fgColor="E8F5E0")
    title_font  = Font(bold=True, color="FFFFFF", size=14)
    bold        = Font(bold=True, size=11)
    thin        = Side(style="thin", color="DDDDDD")
    border      = Border(left=thin, right=thin, top=thin, bottom=thin)
    center      = Alignment(horizontal="center", vertical="center")
    left_wrap   = Alignment(horizontal="left", vertical="center", wrap_text=True)

    headers = [
        "Наименование", "Артикул", "Масса, г", "Себест., ₸",
        "Белки (порц)", "Жиры (порц)", "Углев. (порц)", "Ккал (порц)", "КДж (порц)",
        "Белки/100г", "Жиры/100г", "Углев./100г", "Ккал/100г",
        "Аллергены",
    ]
    ncols = len(headers)

    def num(val):
        if val is None:
            return None
        try:
            return round(float(val), 2)
        except (TypeError, ValueError):
            return None

    # Заголовок подгруппы
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    c = ws.cell(row=1, column=1, value=f"Блюда на замену: {group.name}")
    c.font = title_font
    c.fill = green_dark
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 26

    # Шапка таблицы
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=2, column=col, value=h)
        c.font = bold
        c.fill = green_light
        c.alignment = center
        c.border = border

    row = 3
    for it in items:
        p = it.product
        weight_g = num(p.net_weight * 1000) if p.net_weight is not None else None
        allergens = ", ".join(a.name for a in p.allergens.all())
        values = [
            p.name, p.article or "", weight_g, num(p.cost),
            num(p.protein_per_serving), num(p.fat_per_serving),
            num(p.carbs_per_serving), num(p.kcal_per_serving), num(p.kj_per_serving),
            num(p.protein), num(p.fat), num(p.carbs), num(p.kcal_per_100),
            allergens,
        ]
        for col, val in enumerate(values, start=1):
            c = ws.cell(row=row, column=col, value=val)
            c.border = border
            if col in (1, 14):
                c.alignment = left_wrap
            else:
                c.alignment = center
        row += 1

    widths = [34, 12, 9, 10, 11, 10, 12, 11, 11, 11, 10, 12, 10, 24]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A3"

    safe_name = "".join(ch for ch in group.name if ch.isalnum() or ch in " -_").strip() or "group"
    filename = f"Замена_{safe_name}.xlsx"

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f"attachment; filename*=UTF-8''{quote(filename)}"
    wb.save(response)
    return response


# ── Рационы Claude (генерация через Anthropic API) ────────────────────────────

def _resolve_ration_proposal(proposal):
    """Разбирает ответ Claude в данные для превью: блюда, суммы КБЖУ, нормы, флаги."""
    all_ids = [
        did
        for meal in proposal.get("meals", [])
        for did in meal.get("dish_ids", [])
    ]
    by_id = {p.pk: p for p in Product.objects.filter(pk__in=all_ids)}

    kcal_cat = proposal.get("kcal_category")
    norm = CalorieCategory.norm_for(kcal_cat)

    meals = []
    tot_kcal = tot_p = tot_f = tot_c = tot_cost = 0
    for meal in proposal.get("meals", []):
        dishes = []
        for did in meal.get("dish_ids", []):
            p = by_id.get(did)
            if not p:
                continue
            kcal = float(p.kcal_per_serving or 0)
            prot = float(p.protein_per_serving or 0)
            fat  = float(p.fat_per_serving or 0)
            carb = float(p.carbs_per_serving or 0)
            cost = float(p.cost or 0)
            tot_kcal += kcal; tot_p += prot; tot_f += fat; tot_c += carb; tot_cost += cost
            dishes.append({
                "id": p.pk,
                "name": p.name, "article": p.article or "",
                "kcal": round(kcal), "protein": round(prot, 1),
                "fat": round(fat, 1), "carbs": round(carb, 1),
                "cost": round(cost),
                "photo": p.photo.url if p.photo else "",
            })
        meals.append({"name": meal.get("meal_name", ""), "dishes": dishes})

    def _flag(val, lo, hi):
        return None if not norm else (val < lo or val > hi)

    return {
        "kcal_category": kcal_cat,
        "meals": meals,
        "reasoning": proposal.get("reasoning", ""),
        "norm": norm,
        "totals": {
            "kcal": round(tot_kcal), "protein": round(tot_p, 1),
            "fat": round(tot_f, 1), "carbs": round(tot_c, 1),
            "cost": round(tot_cost),
        },
        "flags": {
            "kcal":    _flag(tot_kcal, norm.kcal_min, norm.kcal_max) if norm else None,
            "protein": _flag(tot_p, norm.protein_min, norm.protein_max) if norm else None,
            "fat":     _flag(tot_f, norm.fat_min, norm.fat_max) if norm else None,
            "carbs":   _flag(tot_c, norm.carbs_min, norm.carbs_max) if norm else None,
        },
    }


def _ration_fixed_meal_times(ration):
    """Фиксированные приёмы пищи рациона — из его калоража
    (fallback: текущие приёмы пищи рациона). Список MealTime в нужном порядке.
    Структура приёмов пищи задаётся калоражем и не меняется Claude."""
    result, seen = [], set()
    category = CalorieCategory.norm_for(ration.kcal_category)
    if category:
        for meal_time in category.meal_times():
            if meal_time.pk not in seen:
                seen.add(meal_time.pk)
                result.append(meal_time)
    if result:
        return result
    for s in ration.slots.select_related("meal_time").order_by("order", "meal_time__order"):
        if s.meal_time_id and s.meal_time_id not in seen:
            seen.add(s.meal_time_id)
            result.append(s.meal_time)
    return result


def _claude_generate(request, ration):
    """Просит Claude собрать рацион под калорийность, пожелания и ФИКСИРОВАННЫЕ
    приёмы пищи рациона; результат — в сессию."""
    from .claude_rations import generate_ration

    wishes = request.POST.get("wishes", "").strip()
    ration.wishes = wishes or None
    ration.save(update_fields=["wishes"])

    # Каталог без блюд, занятых в других рационах Claude той же калорийности этой группы
    occupied_ids = set()
    if ration.group_id:
        occupied_ids = set(
            ClaudeRationSlot.objects.filter(
                ration__group_id=ration.group_id,
                ration__kcal_category=ration.kcal_category,
                product__isnull=False,
            ).exclude(ration=ration).values_list("product_id", flat=True)
        )
    products = (
        Product.objects.prefetch_related("meal_categories")
        .exclude(pk__in=occupied_ids).order_by("name")
    )
    norms = CalorieCategory.objects.filter(kcal=ration.kcal_category)
    meal_names = [mt.name for mt in _ration_fixed_meal_times(ration)]

    full_wishes = (
        f"Категория калорийности строго {ration.kcal_category} ккал.\n"
        + (wishes or "(без особых пожеланий)")
    )
    try:
        proposal = generate_ration(full_wishes, products, norms, meal_times=meal_names)
        request.session[f"claude_proposal_{ration.pk}"] = proposal
        request.session.pop(f"claude_error_{ration.pk}", None)
    except Exception as e:
        # Логируем со стеком: раньше ошибка была видна только пользователю
        # в интерфейсе и не попадала в journalctl.
        logger.exception("claude: сборка рациона #%s провалилась: %s", ration.pk, e)
        request.session[f"claude_error_{ration.pk}"] = str(e)
        request.session.pop(f"claude_proposal_{ration.pk}", None)


def _apply_proposal_to_ration(ration, proposal):
    """Наполняет ФИКСИРОВАННЫЕ приёмы пищи рациона блюдами от Claude.

    Структура приёмов пищи задаётся настройками (шаблоном) и НЕ меняется:
    все фиксированные приёмы всегда сохраняются — даже те, что Claude не заполнил
    (для них остаётся пустой слот, чтобы приём не исчез). Общих справочников
    (MealTime) не создаём — только читаем. Повторов блюд в рационе не допускаем.
    (Повторы между рационами одной калорийности группы уже отсечены при генерации.)
    """
    fixed_mts = _ration_fixed_meal_times(ration)

    # Предложение Claude: имя приёма пищи (lower) → список id блюд
    by_meal = {}
    for meal in proposal.get("meals", []):
        name = (meal.get("meal_name") or "").strip().lower()
        by_meal.setdefault(name, []).extend(meal.get("dish_ids", []))

    all_ids = [did for ids in by_meal.values() for did in ids]
    prods = {
        p.pk: p
        for p in Product.objects.filter(pk__in=all_ids).prefetch_related("meal_categories")
    }

    ration.slots.all().delete()
    order = 0
    used_product_ids = set()  # без повторяющихся блюд в одном рационе
    for mt in fixed_mts:
        key = mt.name.strip().lower()
        dish_ids = by_meal.get(key)
        if dish_ids is None:  # запасное частичное совпадение имени
            dish_ids = next(
                (v for k, v in by_meal.items() if k and (k in key or key in k)),
                [],
            )
        placed = 0
        for did in dish_ids:
            p = prods.get(did)
            if not p or p.pk in used_product_ids:  # пропускаем повтор блюда
                continue
            used_product_ids.add(p.pk)
            cats = list(p.meal_categories.all())
            ClaudeRationSlot.objects.create(
                ration=ration,
                meal_time=mt,
                slot_type=cats[0].key if cats else None,
                product=p,
                order=order,
            )
            order += 1
            placed += 1
        if placed == 0:  # Claude не заполнил приём — оставляем пустой слот, чтобы не исчез
            ClaudeRationSlot.objects.create(ration=ration, meal_time=mt, order=order)
            order += 1


def _claude_apply(request, ration):
    proposal = request.session.get(f"claude_proposal_{ration.pk}")
    if proposal:
        _apply_proposal_to_ration(ration, proposal)
    request.session.pop(f"claude_proposal_{ration.pk}", None)
    request.session.pop(f"claude_error_{ration.pk}", None)


# ── Вкладка «Рационы Claude»: группы ──────────────────────────────────────────

def claude_group_list(request):
    """Лендинг вкладки: список групп рационов Claude (аналог ration_group_list)."""
    groups = ClaudeRationGroup.objects.prefetch_related("rations__slots__product").all()
    groups_data = []
    for g in groups:
        rations = list(g.rations.all())
        rations_info = []
        group_cost = 0
        for r in rations:
            slots = list(r.slots.select_related("product", "meal_time").order_by("order", "meal_time__order"))
            filled_slots = [s for s in slots if s.product_id]
            total_kcal = sum(float(s.product.kcal_per_serving or 0) for s in filled_slots)
            total_cost = sum(float(s.product.cost or 0) for s in filled_slots)
            group_cost += total_cost
            rations_info.append({
                "ration": r,
                "slots": filled_slots,
                "total_kcal": round(total_kcal, 1),
                "total_cost": round(total_cost, 2),
                "filled_count": len(filled_slots),
            })
        groups_data.append({
            "group": g,
            "rations_info": rations_info,
            "count": len(rations),
            "group_cost": round(group_cost, 2),
        })

    return render(request, "myapp/claude_group_list.html", {
        "groups_data": groups_data,
        # для модалки заявочного листа: группы замен выбираются наравне с рационами
        "swap_groups": SwapGroup.objects.prefetch_related("items").all(),
        "ration_kcal_choices": CalorieCategory.choices(),
        "slot_icons": SLOT_ICONS,
        "tmpl_data_json": json.dumps(_calorie_meals_preview(), ensure_ascii=False),
        "slot_labels_json": json.dumps(SLOT_LABELS, ensure_ascii=False),
    })


def claude_group_create(request):
    if request.method == "POST":
        name = request.POST.get("name", "").strip()
        description = request.POST.get("description", "").strip()
        if name:
            ClaudeRationGroup.objects.create(name=name, description=description or None)
    return redirect("claude_rations")


def claude_group_edit(request, group_pk):
    group = get_object_or_404(ClaudeRationGroup, pk=group_pk)
    if request.method == "POST":
        name = request.POST.get("name", "").strip()
        description = request.POST.get("description", "").strip()
        if name:
            group.name = name
            group.description = description or None
            group.save()
    return redirect("claude_rations")


def claude_group_delete(request, group_pk):
    group = get_object_or_404(ClaudeRationGroup, pk=group_pk)
    if request.method == "POST":
        group.delete()
    return redirect("claude_rations")


@require_POST
def claude_ration_reorder(request):
    """Порядок рационов Claude + (опц.) перемещение в другую группу Claude.

    Полный аналог ration_reorder, но работает ТОЛЬКО с моделями Claude —
    две вкладки между собой не пересекаются.
    Тело JSON: {ration_id, target_group_id, ordered_ids: [...]}.
    """
    try:
        data = json.loads(request.body or "{}")
    except (ValueError, TypeError):
        return JsonResponse({"ok": False, "message": "Некорректные данные"}, status=400)

    ration_id       = data.get("ration_id")
    target_group_id = data.get("target_group_id")
    ordered_ids     = data.get("ordered_ids") or []

    ration = get_object_or_404(ClaudeRation, pk=ration_id)
    target_group = get_object_or_404(ClaudeRationGroup, pk=target_group_id)

    # Перемещение в ДРУГУЮ группу — проверяем конфликт по блюдам той же калорийности
    if ration.group_id != target_group.pk:
        moved_product_ids = set(
            ClaudeRationSlot.objects.filter(ration=ration, product__isnull=False)
            .values_list("product_id", flat=True)
        )
        if moved_product_ids:
            conflicts = (
                ClaudeRationSlot.objects.filter(
                    ration__group_id=target_group.pk,
                    ration__kcal_category=ration.kcal_category,
                    product_id__in=moved_product_ids,
                )
                .exclude(ration=ration)
                .select_related("product", "ration")
            )
            conflict_names = sorted({
                f"{c.product.name} (в «{c.ration.name}»)"
                for c in conflicts if c.product_id
            })
            if conflict_names:
                return JsonResponse({
                    "ok": False,
                    "conflict": True,
                    "message": (
                        f"Нельзя переместить: в группе «{target_group.name}» "
                        f"уже есть рацион {ration.kcal_category} ккал с теми же блюдами:\n"
                        + "\n".join("• " + n for n in conflict_names)
                    ),
                }, status=409)

        ration.group = target_group
        ration.save(update_fields=["group"])

    # Обновляем порядок всех рационов целевой группы
    for index, rid in enumerate(ordered_ids):
        ClaudeRation.objects.filter(pk=rid, group_id=target_group.pk).update(order=index)

    return JsonResponse({"ok": True})


def claude_ration_create(request, group_pk):
    group = get_object_or_404(ClaudeRationGroup, pk=group_pk)
    if request.method == "POST":
        name = request.POST.get("name", "").strip()
        kcal_category = request.POST.get("kcal_category", "")
        wishes = request.POST.get("wishes", "").strip()
        if name and kcal_category:
            ration = ClaudeRation.objects.create(
                group=group, name=name,
                kcal_category=int(kcal_category),
                wishes=wishes or None,
            )
            # Слоты по приёмам пищи калоража (общий справочник CalorieCategory)
            _seed_slots_from_calorie_category(ration, int(kcal_category), ClaudeRationSlot)
            return redirect("claude_ration_edit", pk=ration.pk)
    return redirect("claude_rations")


def claude_ration_delete(request, pk):
    ration = get_object_or_404(ClaudeRation, pk=pk)
    if request.method == "POST":
        ration.delete()
    return redirect("claude_rations")


def claude_ration_edit(request, pk):
    """Полный редактор рациона Claude: слоты + пикер блюд + панель сборки Claude."""
    ration = get_object_or_404(ClaudeRation, pk=pk)

    if request.method == "POST":
        action = request.POST.get("action", "")

        if action == "update_meta":
            ration.name = request.POST.get("name", ration.name).strip()
            ration.kcal_category = int(request.POST.get("kcal_category", ration.kcal_category))
            ration.notes = request.POST.get("notes", "").strip() or None
            ration.save()

        elif action == "add_slot_with_category":
            meal_time_id = request.POST.get("meal_time_id")
            slot_type    = request.POST.get("slot_type") or None
            if meal_time_id:
                last_order = ration.slots.count()
                ClaudeRationSlot.objects.create(
                    ration=ration,
                    meal_time_id=meal_time_id,
                    slot_type=slot_type,
                    order=last_order,
                )

        elif action == "update_slot_category":
            slot_id = request.POST.get("slot_id")
            slot_type = request.POST.get("slot_type") or None
            try:
                slot = ClaudeRationSlot.objects.get(pk=slot_id, ration=ration)
                slot.slot_type = slot_type
                slot.product = None
                slot.save()
            except ClaudeRationSlot.DoesNotExist:
                pass

        elif action == "update_slot":
            slot_id = request.POST.get("slot_id")
            product_id = request.POST.get("product_id") or None
            try:
                slot = ClaudeRationSlot.objects.get(pk=slot_id, ration=ration)
                slot.product_id = product_id
                slot.save()
            except ClaudeRationSlot.DoesNotExist:
                pass

        elif action == "delete_slot":
            slot_id = request.POST.get("slot_id")
            ClaudeRationSlot.objects.filter(pk=slot_id, ration=ration).delete()

        elif action == "claude_generate":
            _claude_generate(request, ration)

        elif action == "claude_apply":
            _claude_apply(request, ration)

        elif action == "claude_discard":
            request.session.pop(f"claude_proposal_{ration.pk}", None)
            request.session.pop(f"claude_error_{ration.pk}", None)

        return redirect("claude_ration_edit", pk=pk)

    # GET
    slots = list(
        ration.slots.select_related("product", "meal_time").order_by("order", "meal_time__order")
    )

    meal_time_groups = {}
    total_kcal = total_protein = total_fat = total_carbs = total_cost = 0
    for slot in slots:
        mt_id = slot.meal_time_id
        if mt_id not in meal_time_groups:
            meal_time_groups[mt_id] = {"meal_time": slot.meal_time, "slots": []}
        p = slot.product
        kcal    = float(p.kcal_per_serving or 0) if p else 0
        protein = float(p.protein_per_serving or 0) if p else 0
        fat     = float(p.fat_per_serving or 0) if p else 0
        carbs   = float(p.carbs_per_serving or 0) if p else 0
        cost    = float(p.cost or 0) if p else 0
        total_kcal += kcal; total_protein += protein; total_fat += fat
        total_carbs += carbs; total_cost += cost
        meal_time_groups[mt_id]["slots"].append({
            "slot": slot,
            "label": SLOT_LABELS.get(slot.slot_type, slot.slot_type) if slot.slot_type else None,
            "icon":  SLOT_ICONS.get(slot.slot_type, "🍴") if slot.slot_type else "❓",
            "color": SLOT_COLORS.get(slot.slot_type, "green") if slot.slot_type else "green",
            "kcal": round(kcal, 1), "protein": round(protein, 1),
            "fat": round(fat, 1), "carbs": round(carbs, 1),
        })

    occupied_ids = set()
    if ration.group_id:
        occupied_ids = set(
            ClaudeRationSlot.objects.filter(
                ration__group_id=ration.group_id,
                ration__kcal_category=ration.kcal_category,
                product__isnull=False,
            ).exclude(ration=ration).values_list("product_id", flat=True)
        )

    meal_cat_map = {}
    for key, _ in RATION_SLOT_TYPES:
        prods = list(
            Product.objects.filter(meal_categories__key=key)
            .exclude(pk__in=occupied_ids)
            .prefetch_related("allergens").order_by("name")
        )
        meal_cat_map[key] = [_product_picker_dict(p) for p in prods]

    all_meal_times = list(MealTime.objects.order_by("order", "name"))
    all_products = list(
        Product.objects.prefetch_related("meal_categories", "allergens")
        .exclude(pk__in=occupied_ids).order_by("name")
    )
    all_products_json = [_product_picker_dict(p, with_category=True) for p in all_products]
    all_allergens = list(Allergen.objects.order_by("name"))

    norm = CalorieCategory.norm_for(ration.kcal_category)
    norm_flags = {}
    if norm:
        def _out(val, lo, hi):
            return val < lo or val > hi
        norm_flags = {
            "kcal":    _out(total_kcal, norm.kcal_min, norm.kcal_max),
            "protein": _out(total_protein, norm.protein_min, norm.protein_max),
            "fat":     _out(total_fat, norm.fat_min, norm.fat_max),
            "carbs":   _out(total_carbs, norm.carbs_min, norm.carbs_max),
        }

    # Превью предложения Claude (если сгенерировано и ещё не применено)
    claude_proposal = request.session.get(f"claude_proposal_{ration.pk}")
    claude_result = _resolve_ration_proposal(claude_proposal) if claude_proposal else None
    claude_error = request.session.get(f"claude_error_{ration.pk}")

    return render(request, "myapp/claude_ration_edit.html", {
        "ration": ration,
        "claude_result": claude_result,
        "claude_error": claude_error,
        "meal_time_groups": list(meal_time_groups.values()),
        "total_kcal": round(total_kcal, 1),
        "total_protein": round(total_protein, 1),
        "total_fat": round(total_fat, 1),
        "total_carbs": round(total_carbs, 1),
        "total_cost": round(total_cost, 2),
        "norm": norm,
        "norm_flags": norm_flags,
        "slot_types": RATION_SLOT_TYPES,
        "slot_icons": SLOT_ICONS,
        "slot_colors": SLOT_COLORS,
        "slot_labels": SLOT_LABELS,
        "kcal_categories": CalorieCategory.choices(),
        "all_meal_times": all_meal_times,
        "meal_cat_map_json": json.dumps(meal_cat_map, ensure_ascii=False),
        "slot_icons_json": json.dumps(SLOT_ICONS, ensure_ascii=False),
        "slot_labels_json": json.dumps(SLOT_LABELS, ensure_ascii=False),
        "occupied_count": len(occupied_ids),
        "all_products_json": json.dumps(all_products_json, ensure_ascii=False),
        "all_allergens": all_allergens,
    })


# ── Экспорт группы рационов в Excel ───────────────────────────────────────────

def _export_ration_group_xlsx(group):
    """Excel-выгрузка группы рационов. Общий код для обычных рационов и
    рационов Claude — у обеих групп .rations со слотами одинаковой структуры
    (meal_time / slot_type / product), поэтому логика не зависит от модели."""
    import re
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from django.http import HttpResponse
    from urllib.parse import quote

    # Естественная сортировка: "Рацион 2" раньше "Рацион 10".
    # Чистим лишние пробелы, чтобы названия с пробелом в начале не вылезали вперёд.
    def natural_key(ration):
        name = re.sub(r'\s+', ' ', (ration.name or '').strip()).lower()
        parts = re.split(r'(\d+)', name)
        return [int(p) if p.isdigit() else p for p in parts]

    rations = sorted(
        group.rations.prefetch_related("slots__product", "slots__meal_time"),
        key=natural_key,
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Рационы"

    # Стили
    green_dark  = PatternFill("solid", fgColor="2E7D32")
    green_mid   = PatternFill("solid", fgColor="66A03C")
    green_light = PatternFill("solid", fgColor="E8F5E0")
    grey_light  = PatternFill("solid", fgColor="F0F0F0")
    white_bold  = Font(bold=True, color="FFFFFF", size=11)
    title_font  = Font(bold=True, color="FFFFFF", size=14)
    bold        = Font(bold=True, size=11)
    muted       = Font(color="888888", size=10)
    thin        = Side(style="thin", color="DDDDDD")
    border      = Border(left=thin, right=thin, top=thin, bottom=thin)
    center      = Alignment(horizontal="center", vertical="center")
    left_wrap   = Alignment(horizontal="left", vertical="center", wrap_text=True)

    headers = [
        "Приём пищи", "Категория блюда", "Наименование", "Артикул",
        "Масса, г", "Себест., ₸",
        "Белки (порц)", "Жиры (порц)", "Углев. (порц)", "Ккал (порц)", "КДж (порц)",
        "Белки/100г", "Жиры/100г", "Углев./100г", "Ккал/100г",
        "Состав",
    ]
    ncols = len(headers)
    COMP_COL = ncols     # последняя колонка — «Состав» (перенос текста)

    def num(val):
        if val is None:
            return None
        try:
            return round(float(val), 2)
        except (TypeError, ValueError):
            return None

    row = 1
    # ── Заголовок группы ──
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=f"Группа рационов: {group.name}")
    c.font = title_font
    c.fill = green_dark
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 26
    row += 1

    if group.description:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
        c = ws.cell(row=row, column=1, value=group.description)
        c.font = muted
        row += 1

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=f"Всего рационов: {len(rations)}")
    c.font = muted
    row += 2

    # ── По каждому рациону ──
    for ration in rations:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
        c = ws.cell(row=row, column=1,
                    value=f"🍽 {ration.name}  ·  {ration.kcal_category} ккал")
        c.font = white_bold
        c.fill = green_mid
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[row].height = 20
        row += 1

        # Шапка таблицы
        for col, h in enumerate(headers, start=1):
            c = ws.cell(row=row, column=col, value=h)
            c.font = bold
            c.fill = green_light
            c.alignment = center
            c.border = border
        row += 1

        all_slots = sorted(
            ration.slots.all(),
            key=lambda s: (s.meal_time.order if s.meal_time_id else 999, s.order),
        )
        # Приёмы пищи, где уже есть выбранное блюдо
        filled_meal_times = {s.meal_time_id for s in all_slots if s.product_id}
        # Убираем пустые слоты-заглушки (без блюда и без категории),
        # если в этом приёме пищи уже есть блюда
        slots = [
            s for s in all_slots
            if not (s.product_id is None and s.slot_type is None
                    and s.meal_time_id in filled_meal_times)
        ]

        tot_kcal = tot_p = tot_f = tot_c = tot_cost = 0
        for slot in slots:
            p = slot.product
            meal = slot.meal_time.name if slot.meal_time_id else "—"
            cat  = SLOT_LABELS.get(slot.slot_type, slot.slot_type) if slot.slot_type else "—"
            if p:
                weight_g = num(p.net_weight * 1000) if p.net_weight is not None else None
                values = [
                    meal, cat, p.name, p.article or "",
                    weight_g, num(p.cost),
                    num(p.protein_per_serving), num(p.fat_per_serving),
                    num(p.carbs_per_serving), num(p.kcal_per_serving), num(p.kj_per_serving),
                    num(p.protein), num(p.fat), num(p.carbs), num(p.kcal_per_100),
                    p.composition_clean or "",
                ]
                tot_kcal += float(p.kcal_per_serving or 0)
                tot_p    += float(p.protein_per_serving or 0)
                tot_f    += float(p.fat_per_serving or 0)
                tot_c    += float(p.carbs_per_serving or 0)
                tot_cost += float(p.cost or 0)
            else:
                values = [meal, cat, "— блюдо не выбрано —", ""] + [None] * (ncols - 5) + [""]

            for col, val in enumerate(values, start=1):
                c = ws.cell(row=row, column=col, value=val)
                c.border = border
                if col in (3, COMP_COL):
                    c.alignment = left_wrap
                elif col <= 2:
                    c.alignment = Alignment(horizontal="left", vertical="center")
                else:
                    c.alignment = center
            row += 1

        # Итого по рациону
        c = ws.cell(row=row, column=3, value="ИТОГО по рациону:")
        c.font = bold
        c.alignment = Alignment(horizontal="right", vertical="center")
        totals = {5: None, 6: round(tot_cost, 2),
                  7: round(tot_p, 1), 8: round(tot_f, 1),
                  9: round(tot_c, 1), 10: round(tot_kcal, 1)}
        for col in range(1, ncols + 1):
            c = ws.cell(row=row, column=col)
            c.fill = grey_light
            c.border = border
            if col in totals and totals[col] is not None:
                c.value = totals[col]
                c.font = bold
                c.alignment = center
        row += 2

    # Ширина колонок
    widths = [16, 20, 34, 12, 9, 10, 11, 10, 12, 11, 11, 11, 10, 12, 10, 40]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A1"

    # Имя файла (с поддержкой кириллицы)
    safe_name = "".join(ch for ch in group.name if ch.isalnum() or ch in " -_").strip() or "group"
    filename = f"Рационы_{safe_name}.xlsx"

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f"attachment; filename*=UTF-8''{quote(filename)}"
    wb.save(response)
    return response


# ── Материалы ─────────────────────────────────────────────────────────────────

def material_list(request):
    """Вкладка «Материалы»: упаковка и прочее, что уходит в конец заявочного листа."""
    return render(request, "myapp/material_list.html", {
        "materials": Material.objects.all(),
    })


def _parse_material_form(request):
    name = (request.POST.get("name") or "").strip()
    if not name:
        return None
    return {
        "article": (request.POST.get("article") or "").strip(),
        "name":    name,
        "unit":    (request.POST.get("unit") or "").strip() or "1 шт",
    }


def material_create(request):
    if request.method == "POST":
        fields = _parse_material_form(request)
        if fields:
            Material.objects.create(order=Material.objects.count(), **fields)
    return redirect("material_list")


def material_edit(request, pk):
    material = get_object_or_404(Material, pk=pk)
    if request.method == "POST":
        fields = _parse_material_form(request)
        if fields:
            for name, value in fields.items():
                setattr(material, name, value)
            material.save()
    return redirect("material_list")


def material_delete(request, pk):
    material = get_object_or_404(Material, pk=pk)
    if request.method == "POST":
        material.delete()
    return redirect("material_list")


# ── Заявочный лист ────────────────────────────────────────────────────────────

# Шапка — первые 9 строк ручного файла, повторены дословно
ORDER_SHEET_HEADER = [
    "Информация заказчика", "Дата", "Компания", "Эл.почта", "Конт.тел", "Заполнил (И.Ф.)",
]
ORDER_SHEET_TITLE = 'Заявочный лист "Фуд завод"'


def _build_order_sheet_xlsx(products, with_price=False):
    """Заявочный лист: плоский список блюд, следом — материалы, без заголовков
    секций. Оформление повторяет ручной файл: те же 9 строк шапки, рамки,
    ширины. with_price добавляет «Цена прод.» после «Кратность заказа»."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from django.http import HttpResponse
    from urllib.parse import quote

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Лист1"

    thin      = Side(style="thin", color="000000")
    border    = Border(left=thin, right=thin, top=thin, bottom=thin)
    calibri   = Font(name="Calibri", size=12)
    head_font = Font(name="Calibri", size=12, bold=True)
    # Цвета с альфой (FF…) — как в исходном файле, иначе openpyxl пишет 00…
    grey_head = PatternFill("solid", fgColor="FFD9D9D9")   # строка «Артикул | Наименование…»
    grey_sep  = PatternFill("solid", fgColor="FFA6A6A6")   # разделитель под ней
    center    = Alignment(horizontal="center")

    headers = ["Артикул", "Наименование продукта", "Кратность заказа"]
    widths  = [19.29, 81.29, 23.57]
    if with_price:
        headers.append("Цена прод.")
        widths.append(15.0)
    headers.append("ИТОГО")
    widths.append(17.0)
    ncols = len(headers)
    PRICE_COL = 4 if with_price else None

    def bordered_row(row, fill=None, font=None):
        for col in range(1, ncols + 1):
            c = ws.cell(row=row, column=col)
            c.border = border
            if fill: c.fill = fill
            if font: c.font = font

    def item_row(row, article, name, unit, price=None):
        ws.cell(row=row, column=1, value=article or "")
        ws.cell(row=row, column=2, value=name)
        ws.cell(row=row, column=3, value=unit)
        if PRICE_COL:
            ws.cell(row=row, column=PRICE_COL, value=price)
        bordered_row(row, font=calibri)
        ws.cell(row=row, column=1).number_format = "@"   # артикул как текст, чтобы не съел нули
        ws.cell(row=row, column=1).alignment = center
        ws.cell(row=row, column=3).alignment = center
        if PRICE_COL:
            ws.cell(row=row, column=PRICE_COL).alignment = center
        ws.row_dimensions[row].height = 18

    # ── Строки 1–9: шапка ──
    ws.cell(row=1, column=2, value=ORDER_SHEET_TITLE).font = Font(name="Calibri", size=9)
    for i, label in enumerate(ORDER_SHEET_HEADER, start=2):
        ws.cell(row=i, column=1, value=label).font = Font(name="Calibri", size=9)

    for col, title in enumerate(headers, start=1):
        c = ws.cell(row=8, column=col, value=title)
        c.font = head_font
        c.fill = grey_head
        c.border = border
        if col != 2:
            c.alignment = center
    bordered_row(9, fill=grey_sep)

    for r in range(1, 8):
        ws.row_dimensions[r].height = 16.5
    ws.row_dimensions[8].height = 18
    ws.row_dimensions[9].height = 18

    # ── Блюда, следом материалы — одним списком, без заголовков секций ──
    row = 10
    for p in products:
        item_row(
            row, p.article, p.name,
            f"1 {p.packing}" if p.packing else "1 порц",
            float(p.sale_price) if p.sale_price is not None else None,
        )
        row += 1

    for m in Material.objects.all():
        item_row(row, m.article, m.name, m.unit)
        row += 1

    for index, width in enumerate(widths):
        ws.column_dimensions[get_column_letter(index + 1)].width = width

    filename = f"Заявочный лист ПП {timezone.localdate().strftime('%d.%m.%Y')}.xlsx"
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f"attachment; filename*=UTF-8''{quote(filename)}"
    wb.save(response)
    return response


def claude_order_sheet(request):
    """Собирает заявочный лист по отмеченным рационам Claude и группам замен.

    Блюдо попадает в лист один раз, даже если стоит в нескольких рационах и
    ещё и в заменах. Порядок — по алфавиту, подгруппы не выводятся."""
    ration_ids = request.GET.getlist("ration")
    swap_ids   = request.GET.getlist("swap_group")
    if not ration_ids and not swap_ids:
        return redirect("claude_rations")

    product_ids = set(
        ClaudeRationSlot.objects
        .filter(ration_id__in=ration_ids, product__isnull=False)
        .values_list("product_id", flat=True)
    ) | set(
        SwapItem.objects
        .filter(swap_group_id__in=swap_ids)
        .values_list("product_id", flat=True)
    )

    products = sorted(
        Product.objects.filter(pk__in=product_ids),
        key=lambda p: (p.name or "").strip().lower(),
    )
    return _build_order_sheet_xlsx(products, with_price=bool(request.GET.get("price")))


def claude_group_export(request, group_pk):
    group = get_object_or_404(ClaudeRationGroup, pk=group_pk)
    return _export_ration_group_xlsx(group)


# ── iiko синхронизация ────────────────────────────────────────────────────────

@require_POST
def iiko_sync_view(request):
    from .iiko_sync import sync_products_from_iiko
    from .models import IikoSyncLog

    if not IikoSyncLog.can_sync():
        remaining = IikoSyncLog.remaining_today()
        return JsonResponse({
            "ok": False,
            "message": f"Лимит синхронизаций исчерпан. Осталось попыток сегодня: {remaining}",
            "remaining": remaining,
        }, status=429)

    cloud_api_key    = getattr(settings, "IIKO_CLOUD_API_KEY", "")
    org_id           = getattr(settings, "IIKO_ORG_ID", "")
    external_menu_id = getattr(settings, "IIKO_EXTERNAL_MENU_ID", "")
    server_url       = getattr(settings, "IIKO_SERVER_URL", "")
    server_login     = getattr(settings, "IIKO_SERVER_LOGIN", "")
    server_password  = getattr(settings, "IIKO_SERVER_PASSWORD", "")
    cloud_app_id        = getattr(settings, "IIKO_APP_ID", "")
    cloud_client_secret = getattr(settings, "IIKO_CLIENT_SECRET", "")

    if not cloud_api_key or not org_id or not external_menu_id:
        return JsonResponse({
            "ok": False,
            "message": "Не заданы IIKO_CLOUD_API_KEY, IIKO_ORG_ID или IIKO_EXTERNAL_MENU_ID в settings.py",
        }, status=400)

    try:
        result = sync_products_from_iiko(
            cloud_api_key=cloud_api_key,
            org_id=org_id,
            external_menu_id=external_menu_id,
            server_url=server_url,
            server_login=server_login,
            server_password=server_password,
            cloud_app_id=cloud_app_id,
            cloud_client_secret=cloud_client_secret,
        )

        IikoSyncLog.objects.create(
            created_count=result.get("created", 0),
            updated_count=result.get("updated", 0),
            deleted_count=result.get("deleted", 0),
            has_errors=bool(result.get("errors")),
        )

        remaining = IikoSyncLog.remaining_today()
        parts = []
        if result.get("created"):   parts.append(f"Создано: {result['created']}")
        if result.get("updated"):   parts.append(f"Обновлено: {result['updated']}")
        if result.get("deleted"):   parts.append(f"Удалено: {result['deleted']}")
        if result.get("unchanged"): parts.append(f"Без изменений: {result['unchanged']}")
        if result.get("skipped"):   parts.append(f"Пропущено: {result['skipped']}")
        if result.get("errors"):    parts.append(f"Ошибок: {len(result['errors'])}")
        message = " | ".join(parts) if parts else "Нет изменений"
        ok = not result.get("errors") or result.get("created", 0) + result.get("updated", 0) > 0

        return JsonResponse({"ok": ok, "message": message, "remaining": remaining, "detail": result})

    except Exception as e:
        return JsonResponse({"ok": False, "message": str(e)}, status=500)


@require_GET
def iiko_sync_status(request):
    from .models import IikoSyncLog
    remaining = IikoSyncLog.remaining_today()
    used      = IikoSyncLog.used_today()
    last_sync = IikoSyncLog.last_sync_time()
    return JsonResponse({
        "remaining": remaining,
        "used":      used,
        "limit":     IikoSyncLog.DAILY_LIMIT,
        "last_sync": last_sync.strftime("%d.%m.%Y %H:%M") if last_sync else None,
        "can_sync":  IikoSyncLog.can_sync(),
    })


# ── Просмотр группы ───────────────────────────────────────────────────────────

def _render_group_detail(request, group, back_url_name, back_label, edit_url_name):
    """Страница просмотра группы рационов.
    back_url_name / back_label / edit_url_name задают ссылки под конкретную
    вкладку."""
    rations = list(
        group.rations
        .prefetch_related(
            "slots__product__allergens",
            "slots__meal_time",
        )
        .order_by("created_at")
    )

    rations_data = []
    total_slots  = 0
    filled_slots = 0
    group_cost   = 0

    for r in rations:
        # Используем prefetch — НЕ делаем новый select_related
        slots = list(r.slots.all())
        # Сортируем в Python чтобы не делать новый запрос
        slots.sort(key=lambda s: (s.order, s.meal_time.order if s.meal_time else 0))

        meal_groups_map = {}
        total_kcal = total_protein = total_fat = total_carbs = total_cost = 0
        filled = 0

        for slot in slots:
            if not slot.meal_time_id:
                continue

            mt_id = slot.meal_time_id
            if mt_id not in meal_groups_map:
                meal_groups_map[mt_id] = {
                    "meal_time": slot.meal_time,
                    "slots": [],
                }
            meal_groups_map[mt_id]["slots"].append(slot)

            if slot.product_id and slot.product:
                total_kcal    += float(slot.product.kcal_per_serving or 0)
                total_protein += float(slot.product.protein_per_serving or 0)
                total_fat     += float(slot.product.fat_per_serving or 0)
                total_carbs   += float(slot.product.carbs_per_serving or 0)
                total_cost    += float(slot.product.cost or 0)
                filled += 1

        valid_slots  = [s for s in slots if s.meal_time_id]
        total_slots  += len(valid_slots)
        filled_slots += filled
        group_cost   += total_cost

        rations_data.append({
            "ration":        r,
            "meal_groups":   list(meal_groups_map.values()),
            "total_kcal":    round(total_kcal, 1),
            "total_protein": round(total_protein, 1),
            "total_fat":     round(total_fat, 1),
            "total_carbs":   round(total_carbs, 1),
            "total_cost":    round(total_cost, 2),
            "filled":        filled,
            "total":         len(valid_slots),
        })

    return render(request, "myapp/group_detail.html", {
        "group":        group,
        "rations_data": rations_data,
        "total_slots":  total_slots,
        "filled_slots": filled_slots,
        "group_cost":   round(group_cost, 2),
        "back_url_name": back_url_name,
        "back_label":    back_label,
        "edit_url_name": edit_url_name,
    })


def claude_group_detail(request, group_pk):
    group = get_object_or_404(ClaudeRationGroup, pk=group_pk)
    return _render_group_detail(
        request, group,
        back_url_name="claude_rations", back_label="Рационы Claude",
        edit_url_name="claude_ration_edit",
    )